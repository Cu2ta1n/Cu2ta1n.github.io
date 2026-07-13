#!/usr/bin/env python3
import openpyxl, json, re, urllib.request, urllib.parse, time
from pathlib import Path

PAPERS_DIR = Path("/data2/zmy/papers")
EXCEL_FILE = PAPERS_DIR / "灵巧手论文8周阅读计划.xlsx"
OUTPUT_FILE = PAPERS_DIR / "website" / "data.json"

def search_arxiv(title):
    """Search arxiv for paper by title"""
    if not title or len(title) < 10:
        return None
    skip_keywords = ["Mathematical Introduction", "IPDF", "Eigengrasp"]
    if any(kw in title for kw in skip_keywords):
        return None
    
    # Clean title for search
    search_title = title.replace(":", " ").replace("-", " ").strip()
    query = urllib.parse.quote(search_title[:100])
    url = "http://export.arxiv.org/api/query?search_query=all:" + query + "&max_results=3"
    
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "PaperReader/1.0"})
        resp = urllib.request.urlopen(req, timeout=10)
        xml_data = resp.read().decode("utf-8")
        
        # Extract all entries
        entries = re.findall(r'<entry>(.*?)</entry>', xml_data, re.DOTALL)
        for entry in entries:
            title_match = re.search(r'<title>(.*?)</title>', entry, re.DOTALL)
            id_match = re.search(r'<id>(.*?)</id>', entry)
            if title_match and id_match:
                found_title = title_match.group(1).strip().replace('\n', ' ')
                # Check if titles match (fuzzy)
                title_words = set(re.findall(r'\w+', title.lower()))
                found_words = set(re.findall(r'\w+', found_title.lower()))
                overlap = len(title_words & found_words)
                if overlap >= min(5, len(title_words) * 0.6):
                    arxiv_url = id_match.group(1)
                    # Convert to abs URL
                    if '/abs/' in arxiv_url:
                        return arxiv_url
                    elif '/pdf/' in arxiv_url:
                        return arxiv_url.replace('/pdf/', '/abs/')
                    else:
                        arxiv_id = arxiv_url.split('/')[-1]
                        return "https://arxiv.org/abs/" + arxiv_id
    except Exception as e:
        print(f"  Error: {e}")
    return None

def scan_papers_dir():
    papers_map = {}
    for category_dir in PAPERS_DIR.iterdir():
        if category_dir.is_dir() and not category_dir.name.startswith("."):
            for pdf_file in category_dir.glob("*.pdf"):
                papers_map[pdf_file.stem] = str(pdf_file)
    return papers_map

def main():
    print("=" * 60)
    print("论文阅读计划数据更新")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["8周阅读计划"]

    local_papers = scan_papers_dir()
    print(f"找到 {len(local_papers)} 个 PDF 文件")

    papers = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
        seq, week, category, title, priority, read_method, hours, goal, output, status, local_path, link = row
        if title is None:
            continue
        actual_local_path = None
        for local_title, local_file in local_papers.items():
            title_words = set(re.findall(r'\w+', title.lower()))
            local_words = set(re.findall(r'\w+', local_title.lower()))
            overlap = len(title_words & local_words)
            if overlap >= min(3, len(title_words)):
                actual_local_path = local_file
                break
        paper = {
            "id": i, "seq": seq, "week": week, "category": category,
            "title": title, "priority": priority, "readMethod": read_method,
            "hours": hours, "goal": goal, "output": output,
            "arxivLink": link if link else "", "localPath": actual_local_path
        }
        papers.append(paper)

    # Search arxiv links
    papers_no_link = [p for p in papers if not p["arxivLink"]]
    print(f"搜索 arxiv: {len(papers_no_link)} 篇")
    for i, paper in enumerate(papers_no_link):
        t = paper["title"][:50]
        print(f"  [{i+1}] {t}...")
        link = search_arxiv(paper["title"])
        if link:
            paper["arxivLink"] = link
            print(f"    -> {link}")
        else:
            print(f"    -> not found")
        time.sleep(2)

    ws_overview = wb["总览"]
    overview = {
        "totalPapers": len(papers), "totalWeeks": 8,
        "totalHours": sum(p["hours"] for p in papers),
        "mainLine": "功能先验 -> 抓取生成 -> 闭环执行 -> VLA/触觉扩展",
        "phases": []
    }
    for row in ws_overview.iter_rows(min_row=5, max_row=10, values_only=True):
        if row[3] is not None:
            overview["phases"].append({
                "phase": row[3], "weeks": row[4],
                "coreQuestion": row[5], "suggestedOutput": row[6], "checkpoint": row[7]
            })

    weekly = []
    for row in ws_overview.iter_rows(min_row=20, max_row=27, values_only=True):
        if row[0] is not None:
            weekly.append({"week": row[0], "paperCount": row[1], "estimatedHours": row[2]})

    categories = {}
    for paper in papers:
        cat = paper["category"]
        if cat not in categories:
            categories[cat] = {"count": 0, "hours": 0}
        categories[cat]["count"] += 1
        categories[cat]["hours"] += paper["hours"]

    data = {
        "papers": papers, "overview": overview, "weekly": weekly,
        "categories": categories, "lastUpdated": time.strftime("%Y-%m-%d %H:%M:%S")
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    n_links = sum(1 for p in papers if p["arxivLink"])
    print(f"\nDone! Papers: {len(papers)}, With links: {n_links}")
    print(f"Output: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
